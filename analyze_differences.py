#!/usr/bin/env python3
"""
Export differences between internal and fund data in formatted output.
"""

import pandas as pd
import numpy as np
from pathlib import Path
from datetime import datetime
import sys
import glob

def export_formatted_differences(differences_file=None, fund_alias='pi', output_format='excel'):
    """
    Export differences data in the requested format with specific columns:
    - id (NumeroContrato)
    - internal face value
    - fund face value  
    - internal acquisition value
    - fund acquisition value
    
    Supports Google Sheets, Excel, or CSV output.
    """
    
    # If no specific file provided, find the most recent differences file
    if not differences_file:
        differences_dir = Path("reports/differences")
        if differences_dir.exists():
            # Find most recent differences file for the specified fund
            pattern = f"differences_{fund_alias}_*.csv"
            files = list(differences_dir.glob(pattern))
            if not files:
                # Try alternative patterns
                pattern = f"{fund_alias}_fund_differences.csv"
                files = list(differences_dir.glob(pattern))
            if files:
                differences_file = max(files, key=lambda p: p.stat().st_mtime)
                print(f"Using most recent differences file: {differences_file}")
            else:
                print(f"❌ No differences files found for fund '{fund_alias}' in {differences_dir}")
                return
        else:
            print("❌ Reports directory not found. Run export_differences.py first.")
            return
    else:
        differences_file = Path(differences_file)
    
    if not differences_file.exists():
        print(f"❌ File not found: {differences_file}")
        return
    
    # Load the differences
    diff_df = pd.read_csv(differences_file)
    
    print('=== EXPORTING FORMATTED DIFFERENCES ===')
    print(f'Source file: {differences_file}')
    print(f'Total differing records: {len(diff_df):,}')
    
    # Filter out rows with differences under 0.5 cents (likely rounding errors)
    def is_meaningful_difference(row):
        """Check if the row has meaningful differences (>= 0.5 cents)"""
        meaningful = False
        
        # Check ValorFace difference
        if pd.notna(row.get('ValorFace_Difference')):
            if abs(row['ValorFace_Difference']) >= 0.5:
                meaningful = True
                
        # Check ValorAquisicao difference  
        if pd.notna(row.get('ValorAquisicao_Difference')):
            if abs(row['ValorAquisicao_Difference']) >= 0.5:
                meaningful = True
                
        return meaningful
    
    # Apply filter
    filtered_df = diff_df[diff_df.apply(is_meaningful_difference, axis=1)]
    
    print(f'Records with meaningful differences (>= 0.5 cents): {len(filtered_df):,}')
    print(f'Filtered out small differences: {len(diff_df) - len(filtered_df):,}')
    print()
    
    # Create the formatted output DataFrame with requested columns
    formatted_data = []
    
    for _, row in filtered_df.iterrows():
        record = {
            'id': row['NumeroContrato'],
            'internal_face_value': row.get('ValorFace_Internal', ''),
            'fund_face_value': row.get('ValorFace_Fund', ''),
            'internal_acquisition_value': row.get('ValorAquisicao_Internal', ''),
            'fund_acquisition_value': row.get('ValorAquisicao_Fund', '')
        }
        formatted_data.append(record)
    
    # Create DataFrame
    export_df = pd.DataFrame(formatted_data)
    
    # Create output directory
    output_dir = Path("reports/formatted_exports")
    output_dir.mkdir(parents=True, exist_ok=True)
    
    # Generate timestamp
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    
    # Export based on requested format
    if output_format.lower() == 'google_sheets':
        success = export_to_google_sheets(export_df, fund_alias, timestamp)
        if not success:
            print("❌ Google Sheets export failed, falling back to Excel...")
            output_format = 'excel'
    
    if output_format.lower() == 'excel':
        success = export_to_excel(export_df, fund_alias, timestamp, output_dir)
        if not success:
            print("❌ Excel export failed, falling back to CSV...")
            output_format = 'csv'
    
    if output_format.lower() == 'csv':
        export_to_csv(export_df, fund_alias, timestamp, output_dir)
    
    return export_df

def export_to_google_sheets(df, fund_alias, timestamp):
    """Export to Google Sheets."""
    try:
        import gspread
        from google.oauth2.service_account import Credentials
        
        # Try to authenticate with Google Sheets
        # You'll need to set up credentials - check if they exist
        credentials_file = Path("google_credentials.json")
        if not credentials_file.exists():
            print("❌ Google credentials file not found. Please set up google_credentials.json")
            return False
            
        # Set up the credentials
        scope = ["https://spreadsheets.google.com/feeds", "https://www.googleapis.com/auth/drive"]
        creds = Credentials.from_service_account_file(str(credentials_file), scopes=scope)
        client = gspread.authorize(creds)
        
        # Create or open spreadsheet
        spreadsheet_name = f"Fund_Differences_{fund_alias.upper()}_{timestamp}"
        
        try:
            spreadsheet = client.create(spreadsheet_name)
            print(f"📊 Created new Google Spreadsheet: {spreadsheet_name}")
        except Exception as e:
            print(f"❌ Could not create spreadsheet: {e}")
            return False
        
        # Get the first worksheet and rename it
        worksheet = spreadsheet.sheet1
        worksheet.update_title(fund_alias.upper())
        
        # Convert DataFrame to list of lists for Google Sheets
        data = [df.columns.tolist()] + df.values.tolist()
        
        # Update the worksheet
        worksheet.update('A1', data)
        
        # Make the spreadsheet shareable (optional)
        spreadsheet.share('', perm_type='anyone', role='reader')
        
        print(f"✅ Successfully exported to Google Sheets: {spreadsheet.url}")
        return True
        
    except ImportError:
        print("❌ gspread or google-auth libraries not installed. Install with: pip install gspread google-auth")
        return False
    except Exception as e:
        print(f"❌ Google Sheets export failed: {e}")
        return False

def export_to_excel(df, fund_alias, timestamp, output_dir):
    """Export to Excel file."""
    try:
        import openpyxl
        from openpyxl.utils.dataframe import dataframe_to_rows
        from openpyxl.styles import Font, PatternFill
        
        # Create Excel file
        excel_file = output_dir / f"fund_differences_{fund_alias}_{timestamp}.xlsx"
        
        # Create workbook and worksheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = fund_alias.upper()
        
        # Add data to worksheet
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
        
        # Format the header row
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save the file
        wb.save(excel_file)
        print(f"✅ Successfully exported to Excel: {excel_file}")
        return True
        
    except ImportError:
        print("❌ openpyxl library not installed. Install with: pip install openpyxl")
        return False
    except Exception as e:
        print(f"❌ Excel export failed: {e}")
        return False

def export_to_csv(df, fund_alias, timestamp, output_dir):
    """Export to CSV file."""
    try:
        csv_file = output_dir / f"fund_differences_{fund_alias}_{timestamp}.csv"
        df.to_csv(csv_file, index=False)
        print(f"✅ Successfully exported to CSV: {csv_file}")
        return True
    except Exception as e:
        print(f"❌ CSV export failed: {e}")
        return False


if __name__ == "__main__":
    import argparse
    
    parser = argparse.ArgumentParser(description='Export fund differences in formatted output')
    parser.add_argument('--file', help='Specific differences file to process')
    parser.add_argument('--fund', default='pi', help='Fund alias (default: pi)')
    parser.add_argument('--format', choices=['google_sheets', 'excel', 'csv'], default='excel', 
                       help='Output format (default: excel)')
    
    args = parser.parse_args()
    
    # Export formatted differences
    export_formatted_differences(args.file, args.fund, args.format) 